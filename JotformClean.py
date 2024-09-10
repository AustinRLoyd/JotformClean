import pandas as pd
import numpy as np
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Specify the path to the folder containing Excel files
folder_path = "C:/Users/wgranalyst/Desktop/AutomationFolder/Input"

# List all files in the folder
files = os.listdir(folder_path)

# Filter out only Excel files
excel_files = [file for file in files if file.endswith(".xlsx") or file.endswith(".xls")]

# Initialize an empty list to hold data from each Excel file
combined_data = []

# Function to separate date and time
def separate_date_time(date_time_str):
    date_time_obj = datetime.strptime(date_time_str, "%b %d, %Y %I:%M %p")
    date = date_time_obj.strftime("%m/%d/%Y")
    time = date_time_obj.strftime("%H:%M")
    return date, time

# Process each Excel file
for excel_file in excel_files:
    # Construct the full path to the file
    excel_file_path = os.path.join(folder_path, excel_file)

    # Read the Excel file into a DataFrame
    df = pd.read_excel(excel_file_path)

    # Clears the comment column
    df["Comment/Notes: C. Stud Meter, Entertainment, Promotions, Tournaments"] = ""
    
    # Rename the columns to something more manageable and dropping irrelevant columns
    df.rename(
        columns={
            "Comment/Notes: C. Stud Meter, Entertainment, Promotions, Tournaments": "comments",
            "Select the Casino": "casino",
            "# @ High Stakes Area": "High Stakes",
        },
        inplace=True,
    )

    # Apply function to each row and create new columns for date and time
    date_time_split = df["Date and Time of Count"].apply(separate_date_time)
    df["date"], df["time"] = zip(*date_time_split)

    # Drop the original 'Date and Time of Count' Column
    df.drop(columns=["Date and Time of Count"], inplace=True)

    # Combine 'First Name' and 'Last Name' into a new column 'rep'
    df["rep"] = df["First Name"] + " " + df["Last Name"]

    # Drop the 'First Name' and 'Last Name' columns
    df.drop(columns=["First Name", "Last Name", "Geo Stamp", "Submission Date"], inplace=True)

    # Drop the 'Enter Your Email' and 'Timer' columns if they exist
    for col in ["Enter Your Email", "Timer"]:
        if col in df.columns:
            df.drop(columns=[col], inplace=True)

    # Processing the 'High Stakes' Column
    if "High Stakes" in df.columns:
        df["High Stakes"].fillna(0, inplace=True)
        df["High Stakes"] = df["High Stakes"].astype(int)
        df["comments"] = df["High Stakes"].astype(str) + " @ HIGH STAKES SLOT AREA"
        df.loc[df["comments"] == "0 @ HIGH STAKES SLOT AREA", "comments"] = ""
        df.drop("High Stakes", axis=1, inplace=True)

    # Rearrange column positions based on existence of specific columns
    columns_order = [
        "casino", "date", "time", "rep", "comments", "Small Craps PLAYERS",
        "Small Craps TABLES", "High Craps PLAYERS ($25+)", "High Craps TABLES ($25+)",
        "Small Table PLAYERS", "Small TABLES", "High Table PLAYERS ($25+)", 
        "High TABLES ($25+)", "Small Slots (1¢ 5¢ 10¢ 25¢ 50¢)", "Large Slots ($1 $5 $25 $50+)", 
        "Poker PLAYERS", "Poker TABLES", "Bingo", "Small Baccarat PLAYERS", 
        "Small Baccarat TABLES", "High Baccarat PLAYERS ($25+)", "High Baccarat TABLES ($25+)"
    ]
    df = df[[col for col in columns_order if col in df.columns]]

    # Renaming columns to specific format
    df.rename(
        columns={
            "Small Craps PLAYERS": "craps|players|-1",
            "Small Craps TABLES": "craps|open|-1",
            "High Craps PLAYERS ($25+)": "craps|players|25",
            "High Craps TABLES ($25+)": "craps|open|25",
            "Small Table PLAYERS": "other tables games|players|-1",
            "Small TABLES": "other tables games|open|-1",
            "High Table PLAYERS ($25+)": "other tables games|players|25",
            "High TABLES ($25+)": "other tables games|open|25",
            "Small Slots (1¢ 5¢ 10¢ 25¢ 50¢)": "small slots|players|-1",
            "Large Slots ($1 $5 $25 $50+)": "large slots|players|-1",
            "Poker PLAYERS": "poker|players|-1",
            "Poker TABLES": "poker|open|-1",
            "Bingo": "bingo|players|-1",
            "Small Baccarat PLAYERS": "baccarat|players|-1",
            "Small Baccarat TABLES": "baccarat|open|-1",
            "High Baccarat PLAYERS ($25+)": "baccarat|players|25",
            "High Baccarat TABLES ($25+)": "baccarat|open|25",
        },
        inplace=True,
    )

    # Handling specific value checks and replacing as needed
    df.reset_index(drop=True, inplace=True)
    df = df.loc[:, ~df.columns.duplicated()]
    df.drop_duplicates(inplace=True)

    # Handling specific value checks and replacing as needed
    for column in ["baccarat|players|-1", "baccarat|players|25"]:
        open_column = column.replace("players", "open")
        if column in df.columns and open_column in df.columns:
            mask = (df[column] == 0) & (df[open_column] == 0)
            df.loc[mask, [column, open_column]] = -1

    # Calculating total players and total tables
    df["total_players"] = df[
        [
            "craps|players|-1", "craps|players|25", "other tables games|players|-1", 
            "other tables games|players|25"
        ]
    ].sum(axis=1, skipna=True)
    df["total_tables"] = df[
        [
            "craps|open|-1", "craps|open|25", "other tables games|open|-1", 
            "other tables games|open|25"
        ]
    ].sum(axis=1, skipna=True)
    df["tables_confirmed"] = df["total_players"] / df["total_tables"]

    # Converting 0's to -1 in bingo column
    if "bingo|players|-1" in df.columns:
        df["bingo|players|-1"] = df["bingo|players|-1"].apply(lambda x: -1 if x == 0 else x)

    # Replace NaN values with -1 in specified columns
    columns_to_replace = [
        "craps|players|-1", "craps|open|-1", "craps|players|25", "craps|open|25",
        "other tables games|players|-1", "other tables games|open|-1", 
        "other tables games|players|25", "other tables games|open|25", 
        "small slots|players|-1", "large slots|players|-1", "poker|players|-1", 
        "poker|open|-1", "bingo|players|-1"
    ]
    df[columns_to_replace] = df[columns_to_replace].fillna(-1)

    # Add a column to keep track of the source file
    df["source_file"] = excel_file

    # Append processed DataFrame to the list
    combined_data.append(df)

# Combine all processed DataFrames into one
combined_df = pd.concat(combined_data, ignore_index=True)

# Reorder the columns in the DataFrame
desired_order = [
    "casino", "date", "time", "rep", "comments", "craps|players|-1", "craps|open|-1",
    "craps|players|25", "craps|open|25", "other tables games|players|-1", 
    "other tables games|open|-1", "other tables games|players|25", 
    "other tables games|open|25", "small slots|players|-1", "large slots|players|-1", 
    "poker|players|-1", "poker|open|-1", "bingo|players|-1", "baccarat|players|-1", 
    "baccarat|open|-1", "baccarat|players|25", "baccarat|open|25", "total_players", 
    "total_tables", "tables_confirmed", "source_file"
]
combined_df = combined_df[desired_order]

# Where do you want to save default output
file_path = "C:/Users/wgranalyst/Desktop/AutomationFolder/Output/combined_output.xlsx"

# Export combined DataFrame to Excel with specified file path
combined_df.to_excel(file_path, index=False)

# Load the saved Excel file to apply highlighting
wb = load_workbook(file_path)
ws = wb.active

# Define colors for each unique source file
unique_files = combined_df["source_file"].unique()
colors = [
    "ff754c", "ffa34c", "9bff4c", "4cd1ff", "ff4cae", "8d4cff", "CCCCFF", 
    "FFFFCC", "FFCC00", "CCFFCC"
]
file_colors = {
    file: PatternFill(
        start_color=colors[i % len(colors)],
        end_color=colors[i % len(colors)],
        fill_type="solid",
    )
    for i, file in enumerate(unique_files)
}

# Find the index of the "casino" column in the DataFrame
casino_col_index = (combined_df.columns.get_loc("casino") + 1)

# Apply the color fill only to the 'casino' column based on the source file
for row in range(2, ws.max_row + 1):  # Start from the second row to skip the header
    source_file = combined_df.loc[row - 2, "source_file"]  # Adjusting for Python's zero-based indexing
    if source_file in file_colors:
        ws.cell(row=row, column=casino_col_index).fill = file_colors[source_file]  # Only highlight the 'casino' column

# Additional Conditional Formatting Rules
# Define fill colors for specific criteria
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

# Define column indices based on column names
craps_players_col = combined_df.columns.get_loc("craps|players|-1") + 1
craps_open_col = combined_df.columns.get_loc("craps|open|-1") + 1
poker_players_col = combined_df.columns.get_loc("poker|players|-1") + 1
poker_open_col = combined_df.columns.get_loc("poker|open|-1") + 1
small_slots_col = combined_df.columns.get_loc("small slots|players|-1") + 1
tables_confirmed_col = combined_df.columns.get_loc("tables_confirmed") + 1

# Apply conditional formatting rules
for row in range(2, ws.max_row + 1):  # Start from the second row to skip the header
    # Highlight both craps|players|-1 and craps|open|-1 columns in yellow if both are 0
    if (
        ws.cell(row=row, column=craps_players_col).value == 0
        and ws.cell(row=row, column=craps_open_col).value == 0
    ):
        ws.cell(row=row, column=craps_players_col).fill = yellow_fill
        ws.cell(row=row, column=craps_open_col).fill = yellow_fill

    # Highlight both poker|players|-1 and poker|open|-1 columns in yellow if both are 0
    if (
        ws.cell(row=row, column=poker_players_col).value == 0
        and ws.cell(row=row, column=poker_open_col).value == 0
    ):
        ws.cell(row=row, column=poker_players_col).fill = yellow_fill
        ws.cell(row=row, column=poker_open_col).fill = yellow_fill

    # **New Condition: Highlight both craps|players|25 and craps|open|25 columns in yellow if both are 0**
    craps_players_25_col = combined_df.columns.get_loc("craps|players|25") + 1
    craps_open_25_col = combined_df.columns.get_loc("craps|open|25") + 1

    if (
        ws.cell(row=row, column=craps_players_25_col).value == 0
        and ws.cell(row=row, column=craps_open_25_col).value == 0
    ):
        ws.cell(row=row, column=craps_players_25_col).fill = yellow_fill
        ws.cell(row=row, column=craps_open_25_col).fill = yellow_fill

    # Highlight small slots|players|-1 column in red if the value is less than 100
    if (
        ws.cell(row=row, column=small_slots_col).value is not None
        and ws.cell(row=row, column=small_slots_col).value < 100
    ):
        ws.cell(row=row, column=small_slots_col).fill = red_fill

    # Highlight tables_confirmed column in red if the value is less than 2
    if (
        ws.cell(row=row, column=tables_confirmed_col).value is not None
        and ws.cell(row=row, column=tables_confirmed_col).value < 2
    ):
        ws.cell(row=row, column=tables_confirmed_col).fill = red_fill

# Save the modified workbook with the highlighted cells
highlighted_output_path = "C:/Users/wgranalyst/Desktop/AutomationFolder/Output/combined_output_highlighted.xlsx"
wb.save(highlighted_output_path)

print(f"Conditional formatting applied and saved to: {highlighted_output_path}")
